import matplotlib.pyplot as plt
from openpyxl import Workbook,load_workbook 
wb=load_workbook("EntriesGender.xlsx")
ws = wb.active
def main():
    wanted_sport=input('Which sport do you want to learn its'+
                       ' gender distribution? ')
    wanted_sport_for_list=[]
    index=0
    for row in range(2,48):
        sport=str(ws.cell(row,1).value)
        if sport==wanted_sport:
            for column in range(2,4):
                a=str(ws.cell(row,column).value)
                wanted_sport_for_list.insert(index,a)
                index+=1
    genders=['Women','Men']
    plt.pie(wanted_sport_for_list,labels=genders,colors=('m','b'))
    plt.title(wanted_sport)
    plt.show()
main()
    